import openpyxl

book = openpyxl.Workbook()

print(book.sheetnames)

book.create_sheet('computadores')


computadores_page = book['computadores']
computadores_page.append(['Eletrônica', 'Memória Ram', 'Preço'])
computadores_page.append(['Computador 1' , '8gb Ram', 'R$ 2500'])
computadores_page.append(['Computador 2' , '16gb Ram', 'R$ 5500'])
computadores_page.append(['Computador 3' , '32gb Ram', 'R$ 8500'])


book.save('meus computadores.xlsx')